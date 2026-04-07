"""
Merge all lead sources, dedupe, and push to Google Sheets.

Sources:
  1. Charlotte & Nearby Painting Prospects.xlsx  (CRM — painting + others)
  2. Raleigh NC - Painting (BBB.org) - 4 April.xlsx
  3. bbb_painting_nc.csv
  4. bbb_other_nc.csv

Tabs created:
  - Painting Companies
  - Others
"""

import csv
import pickle
import re
from pathlib import Path

import openpyxl
from google.auth.transport.requests import Request
from googleapiclient.discovery import build

# ── Config ──────────────────────────────────────────────────────────────────
TOKEN_FILE   = Path("/Users/allenenriquez/Desktop/Allen Enriquez/token.pickle")
SPREADSHEET_ID = "1G5ATV3g22TVXdaBHfRTkbXthuvnRQuDbx-eI7bUNNz8"

CHARLOTTE_XLSX = Path.home() / "Downloads" / "Charlotte & Nearby Painting Prospects.xlsx"
RALEIGH_XLSX   = Path.home() / "Downloads" / "Raleigh NC - Painting (BBB.org) - 4 April.xlsx"
BBB_PAINTING   = Path.home() / "Desktop" / "bbb_painting_nc.csv"
BBB_OTHER      = Path.home() / "Desktop" / "bbb_other_nc.csv"

HEADERS = [
    "Business Name", "Owner Name / DM", "Phone", "Email", "Website",
    "City", "Service Areas", "Social Media", "LinkedIn", "BBB Rating",
    "Notes", "Call Outcome", "Date Called", "Called", "Connected",
    "Follow-up Date", "Source", "BBB URL",
]

PAINTING_RE = re.compile(r"\bpaint(ing|er|s)?\b", re.IGNORECASE)


# ── Helpers ──────────────────────────────────────────────────────────────────
def dedup_key(name):
    if not name:
        return ""
    name = re.sub(r"^advertisement:\s*\n?", "", str(name), flags=re.I)
    return re.sub(r"[^a-z0-9]", "", name.lower().strip())


def blank():
    return {h: "" for h in HEADERS}


def city_from_address(addr):
    """Extract city from 'Street, City, ST ZIP' format."""
    if not addr:
        return ""
    parts = [p.strip() for p in addr.split(",")]
    return parts[1] if len(parts) >= 2 else parts[0]


# ── Load sources ─────────────────────────────────────────────────────────────
def load_charlotte():
    rows = []
    wb = openpyxl.load_workbook(CHARLOTTE_XLSX)
    ws = wb.active
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        d = blank()
        d["Business Name"]   = str(r[0]).strip() if r[0] else ""
        d["Owner Name / DM"] = str(r[1]).strip() if r[1] else ""
        d["Phone"]           = str(r[2]).strip() if r[2] else ""
        d["Call Outcome"]    = str(r[3]).strip() if r[3] else ""
        d["Notes"]           = str(r[4]).strip() if r[4] else ""
        d["Date Called"]     = str(r[5]).strip() if r[5] else ""
        d["Email"]           = str(r[6]).strip() if r[6] else ""
        d["Website"]         = str(r[7]).strip() if r[7] else ""
        d["Service Areas"]   = str(r[8]).strip() if r[8] else ""
        d["Social Media"]    = str(r[9]).strip() if r[9] else ""
        d["LinkedIn"]        = str(r[10]).strip() if r[10] else ""
        d["Called"]          = str(r[11]).strip() if r[11] else ""
        d["Connected"]       = str(r[12]).strip() if r[12] else ""
        d["Follow-up Date"]  = str(r[13]).strip() if r[13] else ""
        d["Source"]          = "Charlotte CRM"
        rows.append(d)
    print(f"  Charlotte CRM:   {len(rows)} rows")
    return rows


def load_raleigh():
    rows = []
    wb = openpyxl.load_workbook(RALEIGH_XLSX)
    ws = wb.active
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(c for c in r):
            continue
        d = blank()
        d["Business Name"]   = str(r[0]).strip() if r[0] else ""
        d["Phone"]           = str(r[1]).strip() if r[1] else ""
        d["Owner Name / DM"] = str(r[2]).strip() if r[2] else ""
        d["Website"]         = str(r[3]).strip() if r[3] else ""
        social = " | ".join(filter(None, [
            str(r[4]).strip() if r[4] else "",
            str(r[5]).strip() if r[5] else "",
        ]))
        d["Social Media"]    = social
        d["Email"]           = str(r[6]).strip() if r[6] else ""
        d["City"]            = str(r[7]).strip() if r[7] else ""
        d["Service Areas"]   = str(r[8]).strip() if r[8] else ""
        d["BBB Rating"]      = str(r[9]).strip() if r[9] else ""
        notes_parts = []
        if r[10]: notes_parts.append(f"Reviews: {r[10]}")
        if r[11]: notes_parts.append(str(r[11]))
        d["Notes"]           = " | ".join(notes_parts)
        d["Source"]          = "Raleigh BBB (4 Apr)"
        if d["Business Name"]:
            rows.append(d)
    print(f"  Raleigh file:    {len(rows)} rows")
    return rows


def load_bbb_csv(path, source_label):
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            d = blank()
            d["Business Name"]   = r.get("business_name", "").strip()
            d["Owner Name / DM"] = r.get("owner_name", "").strip()
            d["Phone"]           = r.get("phone", "").strip()
            d["Email"]           = r.get("email", "").strip()
            d["City"]            = city_from_address(r.get("address", ""))
            d["BBB Rating"]      = r.get("bbb_rating", "").strip()
            notes_parts = []
            if r.get("years_in_business"): notes_parts.append(f"Years in biz: {r['years_in_business']}")
            if r.get("entity_type"):       notes_parts.append(r["entity_type"])
            d["Notes"]           = " | ".join(notes_parts)
            d["BBB URL"]         = r.get("profile_url", "").strip()
            d["Source"]          = source_label
            if d["Business Name"]:
                rows.append(d)
    print(f"  {source_label}: {len(rows)} rows")
    return rows


# ── Merge & dedupe ────────────────────────────────────────────────────────────
def merge_and_split():
    charlotte = load_charlotte()
    raleigh   = load_raleigh()
    bbb_paint = load_bbb_csv(BBB_PAINTING, "BBB NC Scrape")
    bbb_other = load_bbb_csv(BBB_OTHER,    "BBB NC Scrape")

    painting_seen = {}
    other_seen    = {}

    def add(d, is_painting):
        k = dedup_key(d["Business Name"])
        if not k:
            return
        bucket = painting_seen if is_painting else other_seen
        if k not in bucket:
            bucket[k] = d

    # Charlotte CRM — classify each row
    for d in charlotte:
        add(d, bool(PAINTING_RE.search(d["Business Name"])))

    # Raleigh — all painting
    for d in raleigh:
        add(d, True)

    # BBB csvs — already split
    for d in bbb_paint:
        add(d, True)
    for d in bbb_other:
        add(d, False)

    painting = list(painting_seen.values())
    other    = list(other_seen.values())
    print(f"\n  → Painting (deduped): {len(painting)}")
    print(f"  → Others  (deduped): {len(other)}")
    return painting, other


# ── Google Sheets ─────────────────────────────────────────────────────────────
def get_service():
    with open(TOKEN_FILE, "rb") as f:
        creds = pickle.load(f)
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
    return build("sheets", "v4", credentials=creds)


def ensure_tabs(service, tab_names):
    """Create tabs if missing, return {name: sheetId}."""
    meta = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
    existing = {s["properties"]["title"]: s["properties"]["sheetId"]
                for s in meta["sheets"]}

    requests = []
    for name in tab_names:
        if name not in existing:
            requests.append({"addSheet": {"properties": {"title": name}}})

    if requests:
        service.spreadsheets().batchUpdate(
            spreadsheetId=SPREADSHEET_ID,
            body={"requests": requests},
        ).execute()
        # Re-fetch after creation
        meta = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
        existing = {s["properties"]["title"]: s["properties"]["sheetId"]
                    for s in meta["sheets"]}

    return existing


def write_tab(service, tab_name, rows):
    """Clear tab and write header + rows."""
    values = [HEADERS] + [[r.get(h, "") for h in HEADERS] for r in rows]
    service.spreadsheets().values().clear(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{tab_name}'",
    ).execute()
    service.spreadsheets().values().update(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{tab_name}'!A1",
        valueInputOption="RAW",
        body={"values": values},
    ).execute()
    print(f"  ✓ '{tab_name}' — {len(rows)} rows written")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("Loading sources...")
    painting, other = merge_and_split()

    print("\nConnecting to Google Sheets...")
    service = get_service()
    ensure_tabs(service, ["Painting Companies", "Others"])

    print("Writing tabs...")
    write_tab(service, "Painting Companies", painting)
    write_tab(service, "Others", other)

    print(f"\nDone. {len(painting) + len(other)} total unique leads pushed.")


if __name__ == "__main__":
    main()
